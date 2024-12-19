import openpyxl
import re


def add_space_to_camel_case(word):
    return re.sub(r"(?<=[a-z])([A-Z])", r" \1", word)


def process_xlsx(file_path, output_path):
    # Charger le fichier Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Liste des colonnes à traiter
    columns_to_check = ["Move 1", "Move 2", "Move 3", "Move 4"]
    column_indices = {}

    # Identifier les colonnes par leur nom
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value in columns_to_check:
            column_indices[col[0].value] = col[0].column

    # Mots à modifier
    words_to_check = [
        "AncientPower",
        "SmokeScreen",
        "BubbleBeam",
        "FeatherDance",
        "ThunderPunch",
        "GrassWhistle",
        "DragonBreath",
        "ViceGrip",
        "DynamicPunch",
        "SmellingSalt",
        "SolarBeam",
        "ExtremeSpeed",
    ]

    # Parcourir les lignes et modifier les colonnes spécifiées
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Sauter l'en-tête
        for col_name, col_index in column_indices.items():
            cell = row[col_index - 1]  # Les index de colonne sont 1-based
            if cell.value in words_to_check:
                cell.value = add_space_to_camel_case(cell.value)

    # Sauvegarder le fichier modifié
    wb.save(output_path)
    print(f"Le fichier a été modifié et sauvegardé sous '{output_path}'.")


# Remplacez par vos chemins de fichier
input_file = "EmeraldBattleFrontierComplete.xlsx"
output_file = "EmeraldBattleFrontierComplete.xlsx"
process_xlsx(input_file, output_file)
